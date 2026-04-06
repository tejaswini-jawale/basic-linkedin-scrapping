import sys
import json
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# ─── CONFIG ───────────────────────────────────────────
LINKEDIN_EMAIL    = "your_email@example.com"     # 👈 Replace with your LinkedIn login
LINKEDIN_PASSWORD = "your_password"            # 👈 Replace with your LinkedIn password

# 👇 Add your own LinkedIn company URLs here
COMPANY_URLS = [
    # "https://www.linkedin.com/company/your-company/about/",
    # "https://www.linkedin.com/company/another-company/about/",
]

OUTPUT_FILE = "companies_data.xlsx"
# ──────────────────────────────────────────────────────


def init_driver():
    """Start Chrome browser"""
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def login(driver, email, password):
    """Log in to LinkedIn"""
    print("🔐 Logging in to LinkedIn...")
    driver.get("https://www.linkedin.com/login")
    time.sleep(2)

    driver.find_element(By.ID, "username").send_keys(email)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()

    time.sleep(4)
    print("✅ Logged in successfully!\n")


def safe_get_text(driver, xpaths):
    """Try multiple XPaths and return first match"""
    for xpath in xpaths:
        try:
            el = WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            text = el.text.strip()
            if text:
                return text
        except:
            continue
    return "N/A"


def scrape_company(driver, url):
    """Scrape one company's about page"""
    driver.get(url)
    time.sleep(random.uniform(4, 7))  # Random delay to avoid detection

    data = {}
    data["LinkedIn URL"] = url

    # Company Name
    data["Company Name"] = safe_get_text(driver, [
        "//h1[contains(@class,'org-top-card-summary__title')]",
        "//h1[contains(@class,'top-card')]",
        "//h1"
    ])

    # Description
    data["Description"] = safe_get_text(driver, [
        "//p[contains(@class,'break-words')]",
        "//section[contains(@class,'about')]//p",
        "//div[contains(@class,'org-about-us-organization-description')]"
    ])

    # Website
    try:
        el = driver.find_element(By.XPATH,
            "//dt[contains(text(),'Website')]/following-sibling::dd//a")
        data["Website"] = el.get_attribute("href")
    except:
        data["Website"] = "N/A"

    # Industry
    data["Industry"] = safe_get_text(driver, [
        "//dt[contains(text(),'Industry')]/following-sibling::dd[1]",
        "//div[contains(@data-test-id,'about-us__industry')]"
    ])

    # Company Size
    data["Company Size"] = safe_get_text(driver, [
        "//dt[contains(text(),'Company size')]/following-sibling::dd[1]",
        "//div[contains(@data-test-id,'about-us__size')]"
    ])

    # Headquarters
    data["Headquarters"] = safe_get_text(driver, [
        "//dt[contains(text(),'Headquarters')]/following-sibling::dd[1]",
        "//div[contains(@data-test-id,'about-us__headquarters')]"
    ])

    # Founded Year
    data["Founded"] = safe_get_text(driver, [
        "//dt[contains(text(),'Founded')]/following-sibling::dd[1]",
        "//div[contains(@data-test-id,'about-us__foundedOn')]"
    ])

    # Specialties
    data["Specialties"] = safe_get_text(driver, [
        "//dt[contains(text(),'Specialties')]/following-sibling::dd[1]",
        "//div[contains(@data-test-id,'about-us__specialties')]"
    ])

    # Followers
    data["Followers"] = safe_get_text(driver, [
        "//span[contains(text(),'followers')]",
        "//p[contains(text(),'followers')]"
    ])

    return data


def save_to_excel(data_list, filename):
    """Save all data to Excel with formatting"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LinkedIn Companies"

    headers = [
        "Company Name", "Description", "Website", "Industry",
        "Company Size", "Headquarters", "Founded", "Specialties",
        "Followers", "LinkedIn URL"
    ]

    # Style headers
    header_fill = PatternFill("solid", fgColor="0A66C2")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 30

    # Write data
    for row_idx, record in enumerate(data_list, start=2):
        for col_idx, key in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(key, "N/A"))

    wb.save(filename)
    print(f"\n✅ Data saved to {filename}")


def main():
    driver = init_driver()
    all_data = []

    try:
        login(driver, LINKEDIN_EMAIL, LINKEDIN_PASSWORD)

        for url in COMPANY_URLS:
            print(f"🔍 Scraping: {url}")
            company_data = scrape_company(driver, url)
            all_data.append(company_data)
            print(f"   ✅ Done: {company_data.get('Company Name')}")
            time.sleep(random.uniform(3, 6))  # Delay between companies

    except Exception as e:
        print(f"❌ Error: {e}")

    finally:
        driver.quit()

    if all_data:
        save_to_excel(all_data, OUTPUT_FILE)
    else:
        print("⚠️ No data was collected.")


if __name__ == "__main__":
    main()