from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests
from bs4 import BeautifulSoup
import time
import random
import re
from urllib.parse import urljoin, urlparse, quote_plus
import json
import os
import datetime
from dotenv import load_dotenv
import mysql.connector

db = mysql.connector.connect(
    host=os.getenv("localhost"),
    user=os.getenv("root"),
    password=os.getenv(""),
    database=os.getenv("sphurtin_org_chart_db")
)


cursor = db.cursor()
# Create results directory if it doesn't exist
RESULTS_DIR = os.path.join(os.path.dirname(__file__), 'results')
os.makedirs(RESULTS_DIR, exist_ok=True)

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app)

# ─── LINKEDIN API CONFIG ──────────────────────────────────────────────────────
LINKEDIN_API_KEY = os.getenv('LINKEDIN_API_KEY', '')  # Set your API key as environment variable
LINKEDIN_API_BASE = 'https://api.linkedin.com/v2'

def save_to_mysql(result):
    try:
        query = """
        INSERT INTO linkedin_profiles (
            linkedin_url, full_name, location, about,
            current_company, education, connections, profile_picture
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            full_name = VALUES(full_name),
            location = VALUES(location),
            about = VALUES(about),
            current_company = VALUES(current_company),
            education = VALUES(education),
            connections = VALUES(connections),
            profile_picture = VALUES(profile_picture)
        """

        values = (
            result.get('LinkedIn URL'),
            result.get('Full Name'),
            result.get('Location'),
            result.get('About'),
            result.get('Current Company'),
            result.get('Education'),
            result.get('Connections'),
            result.get('Profile Picture')
        )

        cursor.execute(query, values)
        db.commit()

        print("✅ Saved to MySQL:", result.get('Full Name'))

    except Exception as e:
        print("❌ DB Error:", e)

def get_linkedin_api_headers():
    """Headers for LinkedIn API requests"""
    return {
        'Authorization': f'Bearer {LINKEDIN_API_KEY}',
        'X-Restli-Protocol-Version': '2.0.0',
        'Content-Type': 'application/json',
    }

# ─── HEADERS ──────────────────────────────────────────────────────────────────

def get_stealth_headers():
    return {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
    }

def get_google_headers():
    return {
        'User-Agent': random.choice([
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        ]),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Referer': 'https://www.google.com/',
    }

def get_random_headers():
    return {
        "User-Agent": random.choice([
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
        ]),
        "Accept-Language": "en-US,en;q=0.9",
        "Connection": "keep-alive"
    }

# ─── LINKEDIN API FUNCTIONS ───────────────────────────────────────────────────

def extract_linkedin_id(url):
    """Extract company or profile ID from LinkedIn URL"""
    if '/company/' in url:
        match = re.search(r'/company/([^/?#]+)', url)
        return match.group(1) if match else None
    elif '/in/' in url:
        match = re.search(r'/in/([^/?#]+)', url)
        return match.group(1) if match else None
    return None

def fetch_company_from_api(company_id):
    """Fetch company data from LinkedIn API as fallback"""
    if not LINKEDIN_API_KEY:
        return {}

    try:
        # Get basic company info
        url = f"{LINKEDIN_API_BASE}/organizations/{company_id}"
        response = requests.get(url, headers=get_linkedin_api_headers(), timeout=10)

        if response.status_code == 200:
            data = response.json()
            return {
                "Company Name": data.get('localizedName', 'N/A'),
                "Tagline": data.get('tagline', 'N/A'),
                "Overview": data.get('localizedDescription', 'N/A'),
                "Website": data.get('localizedWebsite', 'N/A'),
                "Industry": data.get('industry', {}).get('localizedName', 'N/A') if data.get('industry') else 'N/A',
                "Company Size": data.get('staffCountRange', {}).get('localizedName', 'N/A') if data.get('staffCountRange') else 'N/A',
                "Headquarters": f"{data.get('headquarter', {}).get('geographicArea', {}).get('localizedName', 'N/A')}, {data.get('headquarter', {}).get('country', {}).get('localizedName', 'N/A')}" if data.get('headquarter') else 'N/A',
                "Founded": str(data.get('foundedOn', {}).get('year', 'N/A')) if data.get('foundedOn') else 'N/A',
                "Type": data.get('organizationType', 'N/A'),
                "Specialties": ', '.join(data.get('specialties', [])) if data.get('specialties') else 'N/A'
            }
        else:
            print(f"API Error: {response.status_code} - {response.text}")
            return {}
    except Exception as e:
        print(f"API request failed: {e}")
        return {}

def fetch_profile_from_api(profile_id):
    """Fetch specific profile data from LinkedIn API: position, connections, and education"""
    if not LINKEDIN_API_KEY:
        return {}

    try:
        data = {}

        # Get basic profile info including connections
        url = f"{LINKEDIN_API_BASE}/people/{profile_id}"
        params = {
            'projection': '(id,firstName,lastName,headline,publicProfileUrl,profilePicture,vanityName,connections)'
        }
        response = requests.get(url, headers=get_linkedin_api_headers(), params=params, timeout=10)

        if response.status_code == 200:
            profile_data = response.json()
            data["Connections"] = profile_data.get('connections', 'N/A')

        # Get current position
        position_url = f"{LINKEDIN_API_BASE}/people/{profile_id}/positions"
        params = {
            'projection': '(elements*(id,title,company,startDate,endDate))',
            'q': 'owned'
        }
        response = requests.get(position_url, headers=get_linkedin_api_headers(), params=params, timeout=10)

        if response.status_code == 200:
            positions_data = response.json()
            elements = positions_data.get('elements', [])
            if elements:
                # Get the most recent position (current position)
                current_position = elements[0]  # API returns in reverse chronological order
                title = current_position.get('title', 'N/A')
                company_name = current_position.get('company', {}).get('name', 'N/A') if current_position.get('company') else 'N/A'
                if title != 'N/A' and company_name != 'N/A':
                    data["Current Company"] = f"{title} at {company_name}"
                elif title != 'N/A':
                    data["Current Company"] = title
                elif company_name != 'N/A':
                    data["Current Company"] = company_name

        # Get education
        education_url = f"{LINKEDIN_API_BASE}/people/{profile_id}/educations"
        params = {
            'projection': '(elements*(id,schoolName,degreeName,fieldOfStudy,startDate,endDate))',
            'q': 'owned'
        }
        response = requests.get(education_url, headers=get_linkedin_api_headers(), params=params, timeout=10)

        if response.status_code == 200:
            education_data = response.json()
            elements = education_data.get('elements', [])
            if elements:
                # Format education data
                education_list = []
                for edu in elements[:3]:  # Limit to 3 most recent
                    school = edu.get('schoolName', 'N/A')
                    degree = edu.get('degreeName', 'N/A')
                    field = edu.get('fieldOfStudy', 'N/A')
                    start_year = edu.get('startDate', {}).get('year', '') if edu.get('startDate') else ''
                    end_year = edu.get('endDate', {}).get('year', '') if edu.get('endDate') else ''

                    edu_str = school
                    if degree and degree != 'N/A':
                        edu_str += f", {degree}"
                    if field and field != 'N/A':
                        edu_str += f" in {field}"
                    if start_year or end_year:
                        years = f"{start_year}-{end_year}" if start_year and end_year else (start_year or end_year)
                        edu_str += f" ({years})"

                    education_list.append(edu_str)

                data["Education"] = " | ".join(education_list) if education_list else 'N/A'
            else:
                data["Education"] = 'N/A'
        else:
            data["Education"] = 'N/A'

        return data

    except Exception as e:
        print(f"API request failed: {e}")
        return {}

def merge_api_fallback(scraped_data, api_data):
    """Merge scraped data with API fallback for missing fields"""
    merged = scraped_data.copy()

    # Only use API data for fields that are "N/A" in scraped data
    for key, value in api_data.items():
        if key in merged and (merged[key] == "N/A" or not merged[key]) and value != "N/A":
            merged[key] = value
            print(f"   API fallback used for: {key}")

    return merged

def human_delay(min_s=1.5, max_s=10):
    time.sleep(random.uniform(min_s, max_s))


def save_results_to_file(results, operation_type, url=None):
    """Save scraping results to a JSON file and return the file path"""
    try:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        if url:
            # Extract identifier from URL for filename
            if '/in/' in url:
                identifier = url.split('/in/')[-1].split('/')[0].split('?')[0]
                filename = f"profile_{identifier}_{timestamp}.json"
            elif '/company/' in url:
                identifier = url.split('/company/')[-1].split('/')[0].split('?')[0]
                filename = f"company_{identifier}_{timestamp}.json"
            else:
                filename = f"{operation_type}_{timestamp}.json"
        else:
            filename = f"{operation_type}_{timestamp}.json"

        filepath = os.path.join(RESULTS_DIR, filename)

        # Save results to file
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)

        return filepath

    except Exception as e:
        print(f"Error saving results to file: {e}")
        return None


def cleanup_old_files(max_age_days=7):
    """Remove result files older than max_age_days"""
    try:
        cutoff_time = datetime.datetime.now() - datetime.timedelta(days=max_age_days)
        deleted_count = 0

        for filename in os.listdir(RESULTS_DIR):
            if filename.endswith('.json'):
                filepath = os.path.join(RESULTS_DIR, filename)
                file_time = datetime.datetime.fromtimestamp(os.path.getmtime(filepath))

                if file_time < cutoff_time:
                    os.remove(filepath)
                    deleted_count += 1

        return deleted_count

    except Exception as e:
        print(f"Error during cleanup: {e}")
        return 0


def normalize_linkedin_url(url):
    if not url:
        raise ValueError("URL required")
    url = url.strip()
    if not re.match(r'^https?://', url, re.IGNORECASE):
        url = 'https://' + url
    parsed = urlparse(url)
    if not parsed.netloc:
        raise ValueError("Invalid URL")
    clean_path = re.sub(r'(?<!:)//+', '/', parsed.path)
    return f"{parsed.scheme}://{parsed.netloc}{clean_path}".rstrip('/')


# ─── GOOGLE DORK: find employee LinkedIn URLs without login ───────────────────

def google_dork_employees(company_name, company_slug, max_results=25):
    found = {}
    session = requests.Session()
    session.headers.update(get_google_headers())

    queries = []
    if company_name and company_name != "N/A":
        queries.append(f'site:linkedin.com/in "{company_name}"')
        queries.append(f'site:linkedin.com/in "{company_name}" -jobs')
    if company_slug:
        queries.append(f'site:linkedin.com/in inurl:{company_slug}')

    for query in queries:
        if len(found) >= max_results:
            break
        try:
            search_url = f"https://www.google.com/search?q={quote_plus(query)}&num=20&hl=en&gl=us"
            print(f"   Google dork: {query}")
            resp = session.get(search_url, timeout=12)

            if resp.status_code == 429:
                print("   Google rate-limited — stopping dork")
                break
            if resp.status_code != 200:
                print(f"   Google status {resp.status_code}")
                continue

            soup = BeautifulSoup(resp.text, 'html.parser')

            for a in soup.find_all('a', href=True):
                href = a['href']
                if '/url?q=' in href:
                    href = re.sub(r'^.*?/url\?q=', '', href)
                    href = re.sub(r'&.*$', '', href)

                if not re.search(r'linkedin\.com/in/[A-Za-z0-9_%-]+', href):
                    continue
                if any(s in href for s in ['authwall', 'login', '/in/search', 'discover']):
                    continue

                clean_url = re.sub(r'\?.*$', '', href).rstrip('/')
                if not clean_url.startswith('http'):
                    clean_url = 'https://www.linkedin.com' + clean_url
                if clean_url in found:
                    continue

                name, title = "N/A", "N/A"
                parent = a
                for _ in range(7):
                    parent = getattr(parent, 'parent', None)
                    if parent is None:
                        break
                    card_text = parent.get_text(separator='\n', strip=True)
                    lines = [l.strip() for l in card_text.splitlines() if l.strip()]
                    if lines:
                        first = re.sub(r'\s*\|\s*LinkedIn.*$', '', lines[0], flags=re.IGNORECASE).strip()
                        if ' - ' in first or ' \u2013 ' in first:
                            parts = re.split(r' [-\u2013] ', first, maxsplit=1)
                            name  = parts[0].strip()[:80]
                            title = parts[1].strip()[:120] if len(parts) > 1 else "N/A"
                        elif first and len(first) < 80:
                            name = first
                        if name != "N/A":
                            break

                found[clean_url] = {"name": name, "title": title, "url": clean_url}

            print(f"   Running total: {len(found)} profiles")
            human_delay(2.5, 5.0)

        except Exception as e:
            print(f"   Dork error: {str(e)[:100]}")
            continue

    return list(found.values())[:max_results]


# ─── EMPLOYEE SCRAPER ─────────────────────────────────────────────────────────

def scrape_company_employees_enhanced(company_url, session, company_name=""):
    print(f"   Hunting employees for: {company_url}")
    employees = []
    seen_urls  = set()

    def add_employees_from_html(html_text, limit=20):
        added = 0
        for raw in re.findall(
            r'https?://(?:www\.)?linkedin\.com/in/([A-Za-z0-9_%-]{3,80})', html_text
        ):
            full_url = f"https://www.linkedin.com/in/{raw.rstrip('/?')}"
            if full_url not in seen_urls:
                seen_urls.add(full_url)
                employees.append({"name": "N/A", "title": "N/A", "url": full_url})
                added += 1
                if added >= limit:
                    break

        for raw in re.findall(r'(?:href=["\']|"url"\s*:\s*")/in/([A-Za-z0-9_%-]{3,80})', html_text):
            full_url = f"https://www.linkedin.com/in/{raw.rstrip('/?')}"
            if full_url not in seen_urls:
                seen_urls.add(full_url)
                employees.append({"name": "N/A", "title": "N/A", "url": full_url})
                added += 1
                if added >= limit:
                    break
        return added

    # Technique 1: /people/ page
    people_url = company_url.rstrip('/') + '/people/'
    try:
        resp = session.get(people_url, timeout=10, headers=get_stealth_headers())
        final = resp.url.lower()
        blocked = any(b in final for b in ['authwall', 'login', 'signin', 'checkpoint'])

        if not blocked and resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'html.parser')
            for a in soup.find_all('a', href=True):
                href = a['href']
                if '/in/' not in href:
                    continue
                if any(s in href for s in ['authwall', 'login', '/in/search']):
                    continue
                full = urljoin('https://www.linkedin.com', href)
                clean = re.sub(r'\?.*$', '', full).rstrip('/')
                if clean in seen_urls:
                    continue
                seen_urls.add(clean)
                nm = a.get_text(strip=True)
                employees.append({
                    "name":  nm[:80] if nm and len(nm) < 80 else "N/A",
                    "title": "N/A",
                    "url":   clean,
                })
            add_employees_from_html(resp.text, limit=20)
            print(f"   Technique 1 (/people/): {len(employees)} profiles")
        else:
            print("   Technique 1: auth-walled")
    except Exception as e:
        print(f"   Technique 1 error: {str(e)[:80]}")

    # Technique 2: LinkedIn search endpoint
    if len(employees) < 5:
        slug_m = re.search(r'/company/([^/?#]+)', company_url, re.IGNORECASE)
        slug   = slug_m.group(1) if slug_m else ""
        if slug:
            search_url = (
                f"https://www.linkedin.com/search/results/people/"
                f"?currentCompany=%5B%22{slug}%22%5D&origin=COMPANY_PAGE_CURATION"
            )
            try:
                resp = session.get(search_url, timeout=10, headers=get_stealth_headers())
                final = resp.url.lower()
                blocked = any(b in final for b in ['authwall', 'login', 'signin', 'checkpoint'])
                if not blocked and resp.status_code == 200:
                    added = add_employees_from_html(resp.text, limit=20)
                    print(f"   Technique 2 (search): +{added} profiles")
                else:
                    print("   Technique 2: auth-walled")
            except Exception as e:
                print(f"   Technique 2 error: {str(e)[:80]}")

    # Technique 3: Google dork
    if len(employees) < 5:
        slug_m = re.search(r'/company/([^/?#]+)', company_url, re.IGNORECASE)
        slug   = slug_m.group(1) if slug_m else ""
        print(f"   Technique 3: Google dork for '{company_name or slug}'")
        dork_results = google_dork_employees(
            company_name=company_name,
            company_slug=slug,
            max_results=25,
        )
        for emp in dork_results:
            if emp['url'] not in seen_urls:
                seen_urls.add(emp['url'])
                employees.append(emp)
        print(f"   Technique 3 (Google dork): total now {len(employees)} profiles")

    return employees[:30]


def format_employees_for_display(employees):
    if not employees:
        return "N/A"
    lines = []
    for emp in employees:
        name  = emp.get("name",  "N/A") or "N/A"
        title = emp.get("title", "N/A") or "N/A"
        url   = emp.get("url",   "")    or ""
        label = name if name != "N/A" else ""
        if title != "N/A":
            label = f"{label} | {title}".strip(" |")
        if not label:
            label = url.split("/in/")[-1] if "/in/" in url else "Unknown"
        if url:
            lines.append(f"{label}\n{url}")
        else:
            lines.append(label)
    return "\n\n".join(lines)


# ─── EMPLOYEE COUNT EXTRACTOR ─────────────────────────────────────────────────

def extract_employee_count(soup_main, soup_about, raw_html_main, raw_html_about):
    """
    Tries multiple strategies to extract the employee count (e.g. '10,001+ employees').
    Returns a string like '10,001+ employees' or 'N/A'.
    """

    # Strategy 0 (HIGHEST PRIORITY): "139,416 associated members" sentence
    # LinkedIn shows this on the about/main page as the real headcount
    for html in [raw_html_about, raw_html_main]:
        # Match: "139,416 associated members" with optional description after
        m = re.search(
            r'([\d,]+)\s+associated members?(?:[^\n<]{0,120})?',
            html, re.IGNORECASE
        )
        if m:
            # Return the full matched sentence, capped at a readable length
            full_match = m.group(0).strip()
            # Clean HTML entities / tags if any leaked through
            full_match = re.sub(r'<[^>]+>', '', full_match).strip()
            full_match = full_match.replace('&amp;', '&').replace('&#39;', "'")
            if len(full_match) > 120:
                full_match = full_match[:120].rsplit(' ', 1)[0] + '…'
            return full_match

        # Also try on plain text of the page
        soup_tmp  = BeautifulSoup(html, 'html.parser')
        full_text = soup_tmp.get_text(separator=' ', strip=True)
        m = re.search(
            r'([\d,]+)\s+associated members?(?:[^\n]{0,120})?',
            full_text, re.IGNORECASE
        )
        if m:
            full_match = m.group(0).strip()
            if len(full_match) > 120:
                full_match = full_match[:120].rsplit(' ', 1)[0] + '…'
            return full_match

    # Strategy 1: dt/dd pairs with 'size' label on both pages
    for soup_target in [soup_about, soup_main]:
        for dt in soup_target.find_all('dt'):
            label = dt.get_text(strip=True).lower()
            if 'size' in label or 'employee' in label:
                dd = dt.find_next_sibling('dd')
                if dd:
                    val = dd.get_text(strip=True)
                    if val:
                        return val

    # Strategy 2: regex on full page text
    for html in [raw_html_about, raw_html_main]:
        soup_tmp = BeautifulSoup(html, 'html.parser')
        full_text = soup_tmp.get_text(separator=' ', strip=True)
        match = re.search(
            r'([\d,]+\+?\s*(?:to|-)\s*[\d,]+\+?|[\d,]+\+?)\s*employees?',
            full_text, re.IGNORECASE
        )
        if match:
            return match.group(0).strip()

    # Strategy 3: JSON blobs in the page
    for html in [raw_html_about, raw_html_main]:
        patterns = [
            r'"staffCount"\s*:\s*(\d+)',
            r'"employeeCount"\s*:\s*(\d+)',
            r'"numEmployees"\s*:\s*(\d+)',
            r'"headcount"\s*:\s*(\d+)',
        ]
        for pat in patterns:
            m = re.search(pat, html)
            if m:
                return f"{int(m.group(1)):,} employees"

    # Strategy 4: JSON-LD numberOfEmployees
    for soup_target in [soup_about, soup_main]:
        for script in soup_target.find_all('script', type='application/ld+json'):
            try:
                ld = json.loads(script.string or "")
                if isinstance(ld, list):
                    ld = ld[0]
                if ld.get('@type') == 'Organization':
                    count = ld.get('numberOfEmployees', None)
                    if isinstance(count, dict):
                        val = count.get('value')
                        if val:
                            return f"{int(val):,} employees"
                    elif count:
                        return f"{int(count):,} employees"
            except Exception:
                pass

    # Strategy 5: look for common LinkedIn size range strings in raw text
    for html in [raw_html_about, raw_html_main]:
        for pattern in [
            r'(\d{1,3}(?:,\d{3})*\+?\s*(?:-|to)\s*\d{1,3}(?:,\d{3})*\+?)\s*employees?',
            r'((?:Self-employed|1-10|11-50|51-200|201-500|501-1,000|1,001-5,000|5,001-10,000|10,001\+))\s*employees?',
        ]:
            m = re.search(pattern, html, re.IGNORECASE)
            if m:
                return m.group(0).strip()

    return "N/A"


# ─── COMPANY SCRAPER ──────────────────────────────────────────────────────────

def scrape_company(url):
    session = requests.Session()
    session.headers.update(get_stealth_headers())
    data = {"LinkedIn URL": url}

    try:
        print(f"Scraping company: {url}")

        # Main page
        resp = session.get(url, timeout=15)
        resp.raise_for_status()
        raw_html_main = resp.text
        soup = BeautifulSoup(raw_html_main, 'html.parser')

        # Company Name
        for sel in [
            'h1[data-test-id="hero-overlay-title"]',
            '.org-page-top-card-summary__title h1',
            '.org-top-card-module__container h1',
            'h1.top-card-layout__title',
            'h1',
        ]:
            el = soup.select_one(sel)
            if el:
                data["Company Name"] = el.get_text(strip=True)
                break

        # Tagline
        for sel in [
            '[data-test-id="hero-overlay-subtitle"]',
            '.org-top-card-module__tagline',
            '.top-card-layout__headline',
        ]:
            el = soup.select_one(sel)
            if el:
                data["Tagline"] = el.get_text(strip=True)
                break

        # Followers
        m = re.search(r'([\d,.KkMm]+)\s*follower[s]?', soup.get_text(), re.IGNORECASE)
        if m:
            data["Followers"] = m.group(1).strip()

        # About page
        about_url = url.rstrip('/') + '/about/'
        resp_about = session.get(about_url, timeout=15)
        raw_html_about = resp_about.text
        soup_about = BeautifulSoup(raw_html_about, 'html.parser')

        for sel in [
            '[data-test-id="about-us__description"]',
            '.org-about-us-organization-description',
            '.org-about-module__description p',
            '.break-words',
        ]:
            el = soup_about.select_one(sel)
            if el:
                raw = el.get_text(strip=True)
                if len(raw) > 50:
                    data["Overview"] = raw[:500] + ("..." if len(raw) > 500 else "")
                    break

        # Structured dt/dd pairs
        dl_items = soup_about.find_all(['dt', 'dd']) or soup.find_all(['dt', 'dd'])
        i = 0
        while i < len(dl_items):
            if dl_items[i].name == 'dt':
                label = dl_items[i].get_text(strip=True).lower()
                if i + 1 < len(dl_items) and dl_items[i + 1].name == 'dd':
                    value = dl_items[i + 1].get_text(strip=True)
                    if 'website'       in label: data["Website"]      = value
                    elif 'phone'       in label or 'telephone' in label or 'contact' in label:
                        data["Phone"] = value
                    elif 'industry'    in label: data["Industry"]     = value
                    elif 'size'        in label or 'employee' in label:
                        data["Company Size"] = value
                    elif 'headquarter' in label or 'hq' in label or 'location' in label:
                        data["Headquarters"] = value
                    elif 'founded'     in label or 'established' in label:
                        data["Founded"]      = value
                    elif 'type'        in label: data["Type"]         = value
                    elif 'specialt'    in label or 'skills' in label:
                        data["Specialties"]  = value
                    i += 2
                else:
                    i += 1
            else:
                i += 1

        # li fallback
        for li in soup_about.find_all('li'):
            tc = li.get_text(strip=True).lower()
            v  = li.get_text(strip=True)
            def _li(key, *keywords):
                if any(k in tc for k in keywords) and not data.get(key):
                    parts = v.split(':')
                    data[key] = parts[-1].strip() if len(parts) > 1 else v
            _li("Website",      'website')
            _li("Phone",        'phone', 'telephone', 'contact')
            _li("Industry",     'industry')
            _li("Company Size", 'size')
            _li("Headquarters", 'headquarter')
            _li("Founded",      'founded', 'established')
            _li("Type",         'type')
            _li("Specialties",  'specialt', 'skills')

        # Phone regex fallback
        if not data.get("Phone"):
            for pat in [
                r'\+\d{1,3}[\s\-]?\d{1,14}',
                r'\(\d{3}\)[\s\-]?\d{3}[\s\-]?\d{4}',
                r'\d{3}[\s\-]?\d{3}[\s\-]?\d{4}',
            ]:
                m = re.search(pat, soup_about.get_text())
                if m:
                    data["Phone"] = m.group(0)
                    break

        # JSON-LD fallback
        for script in soup.find_all('script', type='application/ld+json'):
            try:
                ld = json.loads(script.string)
                if isinstance(ld, list): ld = ld[0]
                if ld.get('@type') == 'Organization':
                    data.setdefault("Website",  ld.get('url',      'N/A'))
                    data.setdefault("Industry", ld.get('industry', 'N/A'))
                    break
            except Exception:
                pass

        # ── EMPLOYEE COUNT (dedicated extraction) ────────────────────────────
        employee_count = extract_employee_count(soup, soup_about, raw_html_main, raw_html_about)
        data["Employee Count"] = employee_count

        # If Company Size already has a value from dt/dd, prefer it
        if data.get("Company Size") and data["Company Size"] != "N/A":
            data["Employee Count"] = data["Company Size"]

        # ── EMPLOYEES (profile links) ────────────────────────────────────────
        company_name = data.get("Company Name", "")
        try:
            emp_list = scrape_company_employees_enhanced(url, session, company_name)
            data["Employees"] = format_employees_for_display(emp_list)
        except Exception as e:
            print(f"Employee scrape error: {e}")
            data["Employees"] = "N/A"

        # Fill missing fields
        for key in [
            "Company Name", "Tagline", "Overview", "Followers",
            "Website", "Phone", "Industry", "Company Size", "Employee Count",
            "Headquarters", "Founded", "Specialties", "Type", "Employees",
        ]:
            data.setdefault(key, "N/A")

        # ── API FALLBACK for missing data ────────────────────────────────────
        missing_fields = [k for k, v in data.items() if v == "N/A"]
        if missing_fields and LINKEDIN_API_KEY:
            print(f"   Missing fields: {missing_fields} - trying API fallback")
            company_id = extract_linkedin_id(url)
            if company_id:
                api_data = fetch_company_from_api(company_id)
                if api_data:
                    data = merge_api_fallback(data, api_data)

    except requests.RequestException as e:
        raise Exception(f"Network error: {str(e)}")
    except Exception as e:
        raise Exception(f"Scraping failed: {str(e)}")

    human_delay()
    return data


# ─── USER SCRAPER ─────────────────────────────────────────────────────────────

def scrape_user(url):
    session = requests.Session()
    session.headers.update({
        "User-Agent": random.choice([
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        ]),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": "1", "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none", "Cache-Control": "max-age=0",
    })

    data = {
        "LinkedIn URL": url, "Full Name": "N/A",
        "Location": "N/A", "About": "N/A",
        "Current Company": "N/A", "Education": "N/A",
        "Connections": "N/A", "Profile Picture": "N/A",
    }

    try:
        print(f"Scraping user: {url}")
        resp = session.get(url, timeout=20, allow_redirects=True)
        final_url = resp.url.lower()
        if any(x in final_url for x in ["authwall", "login", "checkpoint", "signin"]):
            print("   Auth wall — extracting from meta only")

        raw_html = resp.text
        soup = BeautifulSoup(raw_html, "html.parser")

        og_title = soup.find("meta", property="og:title")
        og_desc  = soup.find("meta", property="og:description")
        og_img   = soup.find("meta", property="og:image")

        if og_title and og_title.get("content"):
            content = re.sub(r"\s*\|\s*LinkedIn.*$", "", og_title["content"], flags=re.IGNORECASE).strip()
            if " - " in content:
                parts = content.split(" - ", 1)
                data["Full Name"] = parts[0].strip()
                data["Current Company"]  = parts[1].strip()
            else:
                data["Full Name"] = content

        if og_desc and og_desc.get("content"):
            desc = og_desc["content"].strip()
            loc_match = re.match(r"^([^·•|]{3,60})\s*[·•|]", desc)
            if loc_match:
                candidate = loc_match.group(1).strip()
                if len(candidate) < 60 and not candidate.endswith("."):
                    data["Location"] = candidate
            if len(desc) > 40:
                data["About"] = desc[:500] + ("..." if len(desc) > 500 else "")

        if og_img and og_img.get("content"):
            data["Profile Picture"] = og_img["content"].strip()

        if data["Full Name"] == "N/A":
            title_tag = soup.find("title")
            if title_tag:
                title_text = re.sub(r"\s*\|\s*LinkedIn.*$", "", title_tag.get_text(strip=True), flags=re.IGNORECASE)
                if " - " in title_text:
                    parts = title_text.split(" - ", 1)
                    data["Full Name"] = parts[0].strip()
                    if data["Current Company"] == "N/A":
                        data["Current Company"] = parts[1].strip()
                elif title_text:
                    data["Full Name"] = title_text.strip()

        if data["Full Name"] == "N/A":
            for sel in [
                'h1.text-heading-xlarge',
                '.pv-top-card--list li.inline.t-24.t-black.t-normal.break-words',
                '.text-heading-xlarge.inline.t-24.t-black.t-normal.break-words',
                '.pv-top-card h1',
            ]:
                el = soup.select_one(sel)
                if el and el.get_text(strip=True):
                    data["Full Name"] = el.get_text(strip=True)
                    break

        if data["Current Company"] == "N/A":
            for sel in [
                '.text-body-medium.break-words',
                '.pv-top-card--list-bullet li',
                '.pv-text-details__left-panel h2',
                '.pv-top-card--experience-list',
            ]:
                el = soup.select_one(sel)
                if el and el.get_text(strip=True):
                    data["Current Company"] = el.get_text(strip=True)
                    break

        if data["Location"] == "N/A":
            for sel in [
                '.text-body-small.inline.t-black--light.break-words',
                '.pv-top-card--list-bullet li',
                '.pv-text-details__left-panel .text-body-small.inline',
                '.top-card-layout__first-subline',
            ]:
                el = soup.select_one(sel)
                if el and el.get_text(strip=True):
                    data["Location"] = el.get_text(strip=True)
                    break

        if data["About"] == "N/A":
            for sel in [
                '#about .pv-shared-text-with-see-more__text',
                '#about .inline-show-more-text span',
                '.pv-about-section .lt-line-clamp__raw-line',
                '.artdeco-card__contents p',
            ]:
                el = soup.select_one(sel)
                if el and el.get_text(strip=True):
                    raw = el.get_text(separator=' ', strip=True)
                    data["About"] = raw[:500] + ("..." if len(raw) > 500 else "")
                    break

        if data["Profile Picture"] == "N/A":
            img = soup.select_one('img.pv-top-card-profile-picture__image, img.profile-photo-edit__preview, img.presence-entity__image')
            if img and img.get("src"):
                data["Profile Picture"] = img.get("src").strip()

        json_patterns = {
            "Full Name":        [r'"firstName"\s*:\s*"([^"]{2,60})".*?"lastName"\s*:\s*"([^"]{2,60})"',
                                 r'"name"\s*:\s*"([A-Z][a-zA-Z\s]{2,60})"'],
            "Location":         [r'"locationName"\s*:\s*"([^"]{2,100})"',
                                 r'"geoLocationName"\s*:\s*"([^"]{2,100})"'],
            "About":            [r'"summary"\s*:\s*"([^"]{20,2000})"'],
            "Current Company":  [
                r'"currentCompany"\s*:\s*\{[^}]*?"name"\s*:\s*"([^\"]{2,100})"',
                r'"companyName"\s*:\s*"([^\"]{2,100})"',
                r'"worksFor"\s*:\s*\{[^}]*?"name"\s*:\s*"([^\"]{2,100})"',
            ],
            "Connections":      [r'"numConnections"\s*:\s*(\d{1,6})'],
        }
        for field, patterns in json_patterns.items():
            if data[field] != "N/A":
                continue
            for pattern in patterns:
                match = re.search(pattern, raw_html)
                if match:
                    if match.lastindex and match.lastindex >= 2:
                        value = f"{match.group(1).strip()} {match.group(2).strip()}"
                    else:
                        value = match.group(1).strip()
                    value = value.encode("utf-8").decode("unicode_escape", errors="replace")
                    value = re.sub(r"\\n", " ", value).strip()
                    if field == "About":
                        value = value[:500] + ("..." if len(value) > 500 else "")
                    data[field] = value
                    break

        for script in soup.find_all("script", type="application/ld+json"):
            try:
                ld = json.loads(script.string or "")
                if isinstance(ld, list):
                    ld = next((x for x in ld if x.get("@type") == "Person"), {})
                if ld.get("@type") != "Person":
                    continue
                if data["Full Name"]       == "N/A": data["Full Name"]       = ld.get("name", "N/A")
                if data["Current Company"] == "N/A":
                    position = ld.get("jobTitle", "")
                    org = ld.get("worksFor", [{}])
                    if isinstance(org, list): org = org[0] if org else {}
                    company = org.get("name", "")
                    if position and company:
                        data["Current Company"] = f"{position} at {company}"
                    elif position:
                        data["Current Company"] = position
                    elif company:
                        data["Current Company"] = company
                if data["Location"] == "N/A":
                    addr = ld.get("address", {})
                    data["Location"] = addr.get("addressLocality", "N/A") if isinstance(addr, dict) else "N/A"
                break
            except Exception:
                pass

        found = sum(1 for v in data.values() if v not in ("N/A", url))
        print(f"   {found}/{len(data)-1} fields extracted")

        # ── API CALLS for specific fields: Connections, Education ──
        if LINKEDIN_API_KEY:
            print("   Fetching Connections and Education from API")
            profile_id = extract_linkedin_id(url)
            if profile_id:
                api_data = fetch_profile_from_api(profile_id)
                if api_data:
                    # Prioritize API data for these specific fields
                    for field in ["Current Company", "Connections", "Education"]:
                        if field in api_data and api_data[field] != "N/A":
                            data[field] = api_data[field]
                            print(f"   API data used for: {field}")

        # ── API FALLBACK for other missing data ──────────────────────────────
        missing_fields = [k for k, v in data.items() if v == "N/A"]
        if missing_fields and LINKEDIN_API_KEY:
            print(f"   Missing fields: {missing_fields} - trying API fallback")
            profile_id = extract_linkedin_id(url)
            if profile_id:
                # For other fields, we could add more API calls here if needed
                pass

    except requests.RequestException as e:
        raise Exception(f"Network error: {e}")
    except Exception as e:
        raise Exception(f"Scraping error: {e}")

    human_delay()
    return data


# ─── EXCEL GENERATOR ──────────────────────────────────────────────────────────

def generate_excel_bytes(data_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LinkedIn Data"

    if data_list and "Company Name" in data_list[0]:
        headers = [
            "Company Name", "Tagline", "Overview", "Followers",
            "Website", "Phone", "Industry", "Company Size", "Employee Count",
            "Employees", "Headquarters", "Founded", "Specialties", "Type", "LinkedIn URL",
        ]
    elif data_list and "Full Name" in data_list[0]:
        headers = [
            "Full Name", "Location", "About",
            "Current Company", "Education",
            "Connections", "Profile Picture", "LinkedIn URL",
        ]
    else:
        all_keys = set()
        for item in data_list:
            all_keys.update(item.keys())
        headers = sorted(all_keys)

    header_fill = PatternFill("solid", fgColor="0A66C2")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    col_widths  = {
        "Overview": 70, "Specialties": 60, "Tagline": 50, "About": 70,
        "Employees": 80, "LinkedIn URL": 45, "Profile Picture": 50,
        "Employee Count": 25,
    }

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = col_widths.get(header, 25)

    for row_idx, record in enumerate(data_list, start=2):
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(key, "N/A"))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ─── API ROUTES ───────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_file("index.html")


@app.route("/scrape-profile", methods=["POST"])
def scrape_profile():
    """Scrape a single LinkedIn profile or company"""
    try:
        body = request.get_json(silent=True) or {}
        input_url = (body.get("url") or "").strip()

        if not input_url:
            return jsonify({"error": "URL is required", "code": "MISSING_URL"}), 400

        try:
            url = normalize_linkedin_url(input_url)
        except ValueError as e:
            return jsonify({"error": str(e), "code": "INVALID_URL"}), 400

        lower = url.lower()
        is_company = bool(re.search(r"linkedin\.com/(company|school|showcase)/", lower))
        is_user = bool(re.search(r"linkedin\.com/(in|pub)/", lower))

        if not (is_company or is_user):
            return jsonify({
                "error": "URL must be a LinkedIn company (/company/) or profile (/in/) URL",
                "code": "INVALID_LINKEDIN_URL"
            }), 400

        result = scrape_company(url) if is_company else scrape_user(url)

        # Save results to file
        # 🔥 ADD THIS
        if not is_company:
            save_to_mysql(result)
        saved_file = save_results_to_file([result], "single_profile", url)

        response_data = {
            "results": [result],
            "saved_to_file": saved_file is not None,
            "file_path": saved_file
        }

        return jsonify(response_data)

    except Exception as e:
        return jsonify({"error": f"Scraping failed: {str(e)}", "code": "SCRAPING_ERROR"}), 500


@app.route("/scrape-bulk", methods=["POST"])
def scrape_bulk():
    """Scrape multiple LinkedIn profiles or companies"""
    try:
        body = request.get_json(silent=True) or {}
        urls = body.get("urls")

        if not isinstance(urls, list) or not urls:
            return jsonify({
                "error": "A non-empty list of URLs is required",
                "code": "INVALID_URLS_LIST"
            }), 400

        if len(urls) > 50:  # Limit bulk requests
            return jsonify({
                "error": "Maximum 50 URLs allowed per request",
                "code": "TOO_MANY_URLS"
            }), 400

        results, errors = [], []

        for raw_url in urls:
            try:
                if not raw_url or not str(raw_url).strip():
                    raise ValueError("Empty URL provided")

                clean_url = normalize_linkedin_url(str(raw_url))
                lower_url = clean_url.lower()
                is_company = bool(re.search(r"linkedin\.com/(company|school|showcase)/", lower_url))
                is_user = bool(re.search(r"linkedin\.com/(in|pub)/", lower_url))

                if not (is_company or is_user):
                    errors.append({
                        "url": raw_url,
                        "error": "Must be a LinkedIn /company/ or /in/ URL",
                        "code": "INVALID_LINKEDIN_URL"
                    })
                    continue

                result = scrape_company(clean_url) if is_company else scrape_user(clean_url)
                results.append(result)

            except Exception as e:
                errors.append({
                    "url": raw_url,
                    "error": str(e),
                    "code": "SCRAPING_ERROR"
                })

        # Save results to file if we have any successful results
        saved_file = None
        if results:
            saved_file = save_results_to_file(results, "bulk_scraping")

        response_data = {
            "results": results,
            "errors": errors,
            "saved_to_file": saved_file is not None,
            "file_path": saved_file,
            "total_processed": len(results) + len(errors),
            "successful": len(results),
            "failed": len(errors)
        }

        return jsonify(response_data)

    except Exception as e:
        return jsonify({"error": f"Request processing failed: {str(e)}", "code": "REQUEST_ERROR"}), 500


@app.route("/download-excel", methods=["POST"])
def download_excel():
    """Export scraped data to Excel format"""
    try:
        data = request.get_json(silent=True)

        if isinstance(data, dict):
            data_list = data.get("results", [])
        elif isinstance(data, list):
            data_list = data
        else:
            return jsonify({
                "error": "Request body must be a JSON object with 'results' array or a JSON array",
                "code": "INVALID_DATA_FORMAT"
            }), 400

        if not data_list:
            return jsonify({
                "error": "No data to export",
                "code": "NO_DATA"
            }), 400

        bio = generate_excel_bytes(data_list)
        return send_file(
            bio,
            as_attachment=True,
            download_name="linkedin_data.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return jsonify({"error": f"Excel generation failed: {str(e)}", "code": "EXCEL_ERROR"}), 500


@app.route("/upload-urls", methods=["POST"])
def upload_urls():
    """Upload a file containing URLs to scrape"""
    try:
        if 'file' not in request.files:
            return jsonify({
                "error": "No file part in the request",
                "code": "NO_FILE"
            }), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({
                "error": "No file selected",
                "code": "EMPTY_FILENAME"
            }), 400

        if file:
            try:
                filename = file.filename.lower()
                if filename.endswith(('.xlsx', '.xls')):
                    # Handle Excel file
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    urls = []
                    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):  # Start from row 2, first column
                        url = row[0]
                        if url and str(url).strip():
                            urls.append(str(url).strip())
                elif filename.endswith(('.txt', '.csv')):
                    # Handle text/CSV file
                    content = file.read().decode('utf-8')
                    urls = [line.strip() for line in content.splitlines() if line.strip()]
                else:
                    return jsonify({
                        "error": "Unsupported file type. Use .xlsx, .xls, .txt, or .csv",
                        "code": "UNSUPPORTED_FILE_TYPE"
                    }), 400

                if not urls:
                    return jsonify({
                        "error": "No URLs found in file",
                        "code": "NO_URLS_IN_FILE"
                    }), 400

                if len(urls) > 50:  # Limit file uploads
                    return jsonify({
                        "error": "Maximum 50 URLs allowed per file",
                        "code": "TOO_MANY_URLS"
                    }), 400

            except Exception as e:
                return jsonify({
                    "error": f"Error reading file: {str(e)}",
                    "code": "FILE_READ_ERROR"
                }), 400

        results, errors = [], []

        for raw_url in urls:
            try:
                if not raw_url or not str(raw_url).strip():
                    raise ValueError("Empty URL in file")

                clean_url = normalize_linkedin_url(str(raw_url))
                lower_url = clean_url.lower()
                is_company = bool(re.search(r"linkedin\.com/(company|school|showcase)/", lower_url))
                is_user = bool(re.search(r"linkedin\.com/(in|pub)/", lower_url))

                if not (is_company or is_user):
                    errors.append({
                        "url": raw_url,
                        "error": "Must be a LinkedIn /company/ or /in/ URL",
                        "code": "INVALID_LINKEDIN_URL"
                    })
                    continue

                result = scrape_company(clean_url) if is_company else scrape_user(clean_url)
                results.append(result)

            except Exception as e:
                errors.append({
                    "url": raw_url,
                    "error": str(e),
                    "code": "SCRAPING_ERROR"
                })

        # Save results to file if we have any successful results
        saved_file = None
        if results:
            saved_file = save_results_to_file(results, "file_upload")

        response_data = {
            "results": results,
            "errors": errors,
            "saved_to_file": saved_file is not None,
            "file_path": saved_file,
            "total_processed": len(results) + len(errors),
            "successful": len(results),
            "failed": len(errors),
            "source_file": file.filename
        }

        return jsonify(response_data)

    except Exception as e:
        return jsonify({"error": f"Request processing failed: {str(e)}", "code": "REQUEST_ERROR"}), 500


@app.route("/health", methods=["GET"])
def health_check():
    """Health check endpoint"""
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.datetime.utcnow().isoformat() + "Z",
        "version": "1.0.0"
    })


@app.route("/status", methods=["GET"])
def api_status():
    """Get API status and configuration"""
    try:
        result_files_count = len([f for f in os.listdir(RESULTS_DIR) if f.endswith('.json')])
    except:
        result_files_count = 0

    return jsonify({
        "status": "running",
        "api_key_configured": bool(LINKEDIN_API_KEY),
        "version": "1.0.0",
        "results_directory": RESULTS_DIR,
        "saved_result_files": result_files_count,
        "endpoints": [
            "/",
            "/health",
            "/status",
            "/version",
            "/api/docs",
            "/results",
            "/scrape-profile",
            "/scrape-bulk",
            "/download-excel",
            "/upload-urls",
            "/cleanup"
        ]
    })


@app.route("/api/docs", methods=["GET"])
def api_docs():
    """Return API documentation"""
    try:
        with open("API_DOCUMENTATION.md", "r", encoding="utf-8") as f:
            docs = f.read()
        response = app.response_class(
            response=docs,
            status=200,
            mimetype="text/markdown"
        )
        response.headers["Access-Control-Allow-Origin"] = "*"
        return response
    except FileNotFoundError:
        return jsonify({"error": "Documentation not found"}), 404


@app.route("/version", methods=["GET"])
def api_version():
    """Get API version information"""
    return jsonify({
        "version": "1.0.0",
        "name": "LinkedIn Scraper API",
        "description": "REST API for scraping LinkedIn profiles and companies with API fallback",
        "endpoints": [
            "GET /",
            "GET /health",
            "GET /status",
            "GET /version",
            "GET /api/docs",
            "GET /results",
            "POST /scrape-profile",
            "POST /scrape-bulk",
            "POST /download-excel",
            "POST /upload-urls",
            "POST /cleanup"
        ]
    })


@app.route("/results", methods=["GET"])
def list_results():
    """List all saved result files"""
    try:
        files = []
        for filename in os.listdir(RESULTS_DIR):
            if filename.endswith('.json'):
                filepath = os.path.join(RESULTS_DIR, filename)
                file_info = {
                    "filename": filename,
                    "filepath": filepath,
                    "size": os.path.getsize(filepath),
                    "created": datetime.datetime.fromtimestamp(os.path.getctime(filepath)).isoformat(),
                    "modified": datetime.datetime.fromtimestamp(os.path.getmtime(filepath)).isoformat()
                }
                files.append(file_info)

        # Sort by creation time (newest first)
        files.sort(key=lambda x: x['created'], reverse=True)

        return jsonify({
            "total_files": len(files),
            "results_directory": RESULTS_DIR,
            "files": files
        })

    except Exception as e:
        return jsonify({"error": f"Failed to list results: {str(e)}", "code": "LIST_ERROR"}), 500


@app.route("/cleanup", methods=["POST"])
def cleanup_results():
    """Clean up old result files"""
    try:
        body = request.get_json(silent=True) or {}
        max_age_days = body.get("max_age_days", 7)  # Default 7 days

        if not isinstance(max_age_days, int) or max_age_days < 1:
            return jsonify({
                "error": "max_age_days must be a positive integer",
                "code": "INVALID_MAX_AGE"
            }), 400

        deleted_count = cleanup_old_files(max_age_days)

        return jsonify({
            "message": f"Cleaned up {deleted_count} old result files",
            "max_age_days": max_age_days,
            "results_directory": RESULTS_DIR
        })

    except Exception as e:
        return jsonify({"error": f"Cleanup failed: {str(e)}", "code": "CLEANUP_ERROR"}), 500


if __name__ == "__main__":
    print("LinkedIn Scraper API running at http://localhost:3000")
    print("Employee Count: extracted from page text, dt/dd pairs, JSON blobs, JSON-LD")
    print("Use responsibly — public data only.")

    import sys
    if len(sys.argv) > 1 and "linkedin.com" in sys.argv[1]:
        url = sys.argv[1]
        try:
            result = scrape_company(url) if "/company/" in url else scrape_user(url)
            for k, v in result.items():
                print(f"  {k}: {str(v)[:120]}")
        except Exception as e:
            print(f"Error: {e}")

    app.run(debug=True, port=3000, host='0.0.0.0')
